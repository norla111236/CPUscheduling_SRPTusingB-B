import pandas as pd
from sumOfC.hw4_dfsSRPT import runDfs
from sumOfC import hw4_bfsSRPT 
from sumOfC import bfsToDfs
import math

all_results=[]

# write the output result to excel file
def writeResult():
    # the dataframe store current results, write the dataframe info to an excel file
    df=pd.DataFrame(all_results,columns=['Number of jobs', 
        'DFS_sumOfC','DFS_elapse time','DFS_nodeVisited','DFS UB_update', 
        'BFS_sumOfC','BFS elapse time','BFS_nodeVisited','BFS UB_update', 
        'bfsToDfs_sumOfC','bfsToDfs elapse time','bfsToDfs_nodeVisited','bfsToDfs UB_update'])

    writer=pd.ExcelWriter('BB_results.xlsx')
    df.to_excel(writer,index=False,sheet_name='B&B report',engine='xlsxwriter')
    writer.save()

    # adjust color format, column fornmat and etc. of the excel file for a clearer appearance
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
    wb = load_workbook(filename = 'BB_results.xlsx')
    ws = wb.active
    font = Font(name='Calibri',size=12)
    titleFont = Font(name='Calibri',size=13,bold=True)
    alignment=Alignment(horizontal='center',vertical='center')
    greenFill = PatternFill(start_color='00CCFFCC', end_color='DAF7A6',fill_type='solid')
    blueFill= PatternFill(start_color='00CCFFFF', end_color='95CCFF',fill_type='solid')
    yelloFill=PatternFill(start_color='00FFFF99', end_color='F9FF95',fill_type='solid')
    thin = Side(border_style="thin", color="000000")
    thick = Side(border_style="thick", color="000000")

    for column in ws['A:M']:
        for cell in column:
            cell.font=font
            cell.alignment=alignment
            cell.border=Border(top=thin, left=thin, right=thin, bottom=thin)

    for cell in ws[1]:
        cell.border=Border(top=thick, left=thin, right=thin, bottom=thick)
        cell.font=titleFont
    for column in ws['A:M']:
        column[len(column)-1].border=Border(top=thin, left=thin, right=thin, bottom=thick)

    for column in ws['B:E']:
        for cell in column:
            cell.fill =blueFill 
                    
    for column in ws['F:I']:
        for cell in column:
            cell.fill =greenFill 
                    
    for column in ws['J:M']:
        for cell in column:
            cell.fill = yelloFill
            
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = max(20,eLen)
    ws.column_dimensions['C'].width = 20      
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 19
    ws.column_dimensions['F'].width = max(20,eLen)
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = max(21,eLen)
    ws.column_dimensions['K'].width = 26
    ws.column_dimensions['L'].width = 26
    ws.column_dimensions['M'].width = 26
    wb.save('BB_results.xlsx')


# start from five jobs, run dfs, bfs, bfsToDfs in every while loop 
# the while loop only stops when  dfs, bfs, bfsToDfs all encounter error
# for dfs, bfs, bfsToDfs, once encounter error, stops and save the error message to the excel file
# saves the current result to excel file in every while loop
job_num=5
dfsStop=False
bfsStop=False
bToDStop=False
eLen=0
while (not dfsStop) or (not bfsStop) or (not bToDStop): 
    dfsResult={'sumOfC':None,'elapseTime':None,'nodeVisited':None,'updateUB':None}
    bfsResult={'sumOfC':None,'elapseTime':None,'nodeVisited':None,'updateUB':None}
    bToDResult={'sumOfC':None,'elapseTime':None,'nodeVisited':None,'updateUB':None}

    if not dfsStop:
        try:
            # if job_num==27:
            #     raise Exception("Sorry, no numbers below zero") 
            dfsResult=runDfs(job_num)
        except Exception as e:
            # print('Exception!!!!:',str(e))
            dfsResult['sumOfC']=str(e)
            dfsStop=True
            eLen=max(len(str(e),eLen))

    if not bfsStop:
        try:
            # if job_num==23:
            #     raise Exception("Sorry, no numbers below zero")
            bfsResult=hw4_bfsSRPT.runBfs(job_num)
        except Exception as e:
            # print('Exception!!!!:',str(e))
            bfsResult['sumOfC']=str(e)
            bfsStop=True
            eLen=max(len(str(e),eLen))
            
    if not bToDStop:
        try:
            # if job_num==7:
            #     raise Exception("Sorry, no numbers below zero")
            bToDResult=bfsToDfs.runBfs(job_num,math.floor(job_num/2))
            # print('bTodResult:',bTodResult)
        except Exception as e:
            # print('Exception!!!!:',str(e))
            bToDResult['sumOfC']=str(e)
            bToDStop=True
            eLen=max(len(str(e),eLen))

    print('\ndfsStop:',dfsStop,'\nbfsStop:',bfsStop,'\nbToDStop:',bToDStop)        
    all_results.append([job_num,
        dfsResult['sumOfC'],dfsResult['elapseTime'],dfsResult['nodeVisited'],dfsResult['updateUB'],
        bfsResult['sumOfC'],bfsResult['elapseTime'],bfsResult['nodeVisited'],bfsResult['updateUB'],
        bToDResult['sumOfC'],bToDResult['elapseTime'],bToDResult['nodeVisited'],bToDResult['updateUB']])
    # if job_num%5==0:
    writeResult()    
    job_num+=1

